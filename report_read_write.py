import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Font


def read_excel(tvc_file: str):
    """Reads following columns from TVC report:
    Email address
    Name
    TVC statistics (for four weeks - last 4 columns)
    Only this data is required for further analysis

    This function create extra list with last 4 columns names,
    which are basically fridays of the weeks that TVC is checked for
    This is used to calculated how many hours must be claimed by the user"""

    wb = openpyxl.load_workbook(tvc_file)
    ws = wb.active

    row_max = ws.max_row
    col_max = ws.max_column

    fridays = [ws.cell(row=1, column=col).value for col in range(col_max, col_max - 4, -1)]

    tvc_data = []

    for r in range(2, row_max):
        if ws.cell(row=r, column=9).value == 'Y':
            tvc_data.append({
                'email_id': ws.cell(row=r, column=2).value,
                'username': ws.cell(row=r, column=1).value,
                ws.cell(row=1, column=col_max - 3).value: ws.cell(row=r, column=col_max - 3).value,
                ws.cell(row=1, column=col_max - 2).value: ws.cell(row=r, column=col_max - 2).value,
                ws.cell(row=1, column=col_max - 1).value: ws.cell(row=r, column=col_max - 1).value,
                ws.cell(row=1, column=col_max).value: ws.cell(row=r, column=col_max).value,
            })

    return tvc_data, fridays

def tvc_archive(report_name, tvc_data):
    """Archive data in the report
    New sheet is created (Archive)"""

    wb_result = openpyxl.load_workbook(report_name)

    if 'Archive' not in wb_result.sheetnames:
        ws_result = wb_result.create_sheet("Archive", 1)
        result_headers = ['Person', 'E-mail', 'Week', 'Hour(s)', 'Minute(s)']

        col_num = 0
        for header in result_headers:
            ws_result.row_dimensions[1].font = Font(name='Calibri', bold=True)
            col_num += 1
            ws_result.cell(row=1, column=col_num).value = header
            ws_result.cell(row=2, column=1).value = datetime.today().strftime("%d-%m-%Y")

        ws_result.column_dimensions['A'].width = 20
        ws_result.column_dimensions['B'].width = 30
        ws_result.column_dimensions['C'].width = 10

        for result in tvc_data:
            row_num = ws_result.max_row + 1
            ws_result.cell(row=row_num, column=1).value = result['email']
            ws_result.cell(row=row_num, column=2).value = result['username']
            ws_result.cell(row=row_num, column=3).value = str(result['week'])
            ws_result.cell(row=row_num, column=4).value = result['hour_miss']
            ws_result.cell(row=row_num, column=5).value = result['min_miss']

            wb_result.save(report_name)
    else:
        print('Data has been already archived\nQuit')
